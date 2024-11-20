import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import io
from fpdf import FPDF
from openpyxl import Workbook
from matplotlib.ticker import MaxNLocator
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Set page title and favicon (logo)
st.set_page_config(page_title="GPA & CGPA Calculator", page_icon="logo.png", layout="wide")

# Function to calculate GPA for a semester, including subjects with 0 credits
def calculate_gpa(marks, credits):
    total_marks = sum([mark * credit for mark, credit in zip(marks, credits) if credit > 0])
    total_credits = sum([credit for credit in credits if credit > 0])
    if total_credits > 0:
        gpa = total_marks / total_credits
    else:
        gpa = 0
    return round(gpa, 2)

# Function to calculate CGPA across all semesters
def calculate_cgpa(gpas, credits):
    total_marks = sum([gpa * credit for gpa, credit in zip(gpas, credits) if credit > 0])
    total_credits = sum([credit for credit in credits if credit > 0])
    if total_credits > 0:
        cgpa = total_marks / total_credits
    else:
        cgpa = 0
    return round(cgpa, 2)

# Function to wrap text for long subject names in PDF
def wrap_text(pdf, text, width):
    words = text.split(' ')
    lines = []
    current_line = []
    for word in words:
        if pdf.get_string_width(' '.join(current_line + [word])) <= width:
            current_line.append(word)
        else:
            lines.append(' '.join(current_line))
            current_line = [word]
    lines.append(' '.join(current_line))
    return lines

# Function to generate PDF report
def generate_pdf(data, cgpa=None):
    pdf_output = io.BytesIO()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="GPA Report", ln=True, align='C')
    pdf.ln(10)

    pdf.cell(20, 10, "S.No", border=1, align='C')
    pdf.cell(60, 10, "Subject Name", border=1, align='C')
    pdf.cell(30, 10, "Marks", border=1, align='C')
    pdf.cell(30, 10, "Credits", border=1, align='C')
    pdf.cell(20, 10, "GPA", border=1, align='C')
    pdf.ln()

    sno = 1
    for semester_data in data:
        pdf.set_font("Arial", style="B", size=12)
        pdf.cell(20, 10, f"Sem {semester_data['semester']}", border=1, align='C')
        pdf.cell(60, 10, "", border=1, align='C')
        pdf.cell(30, 10, "", border=1, align='C')
        pdf.cell(30, 10, "", border=1, align='C')
        pdf.cell(20, 10, f"{semester_data['gpa']:.2f}", border=1, align='C')
        pdf.ln()

        pdf.set_font("Arial", size=10)
        for sub_name, sub_mark, sub_credit in zip(semester_data["subjects"], semester_data["marks"], semester_data["credits"]):
            wrapped_name = wrap_text(pdf, sub_name, 60)
            for i, line in enumerate(wrapped_name):
                pdf.cell(20, 10, str(sno) if i == 0 else "", border=1, align='C')
                pdf.cell(60, 10, line, border=1, align='C')
                pdf.cell(30, 10, str(sub_mark) if i == 0 else "", border=1, align='C')
                pdf.cell(30, 10, str(sub_credit) if i == 0 else "", border=1, align='C')
                pdf.cell(20, 10, "", border=1, align='C')
                pdf.ln()
            sno += 1

    if cgpa is not None:
        pdf.set_font("Arial", style="B", size=12)
        pdf.cell(140, 10, "CGPA", border=1, align='C')
        pdf.cell(20, 10, f"{cgpa:.2f}", border=1, align='C')
        pdf.ln()

    pdf_content = pdf.output(dest='S').encode('latin1')
    pdf_output.write(pdf_content)
    pdf_output.seek(0)
    return pdf_output

# Function to generate Excel report
def generate_excel(data, cgpa=None):
    # Collect all data to structure it for the Excel file
    all_data = []
    for semester_data in data:
        all_data.append({
            "Semester": f"Semester {semester_data['semester']}",
            "Subject": "",
            "Marks": "",
            "Credits": "",
            "GPA": semester_data["gpa"]
        })
        for sub_name, sub_mark, sub_credit in zip(semester_data["subjects"], semester_data["marks"], semester_data["credits"]):
            if sub_credit > 0:  
                all_data.append({
                    "Semester": "",
                    "Subject": sub_name,
                    "Marks": sub_mark,
                    "Credits": sub_credit,
                    "GPA": ""
                })

    if cgpa is not None:
        all_data.append({
            "Semester": "CGPA",
            "Subject": "",
            "Marks": "",
            "Credits": "",
            "GPA": cgpa
        })

    # Create a DataFrame for easier manipulation
    df = pd.DataFrame(all_data)

    # Create an Excel workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = "GPA Report"

    # Define headers
    headers = ["Semester", "Subject", "Marks", "Credits", "GPA"]
    sheet.append(headers)

    # Styling for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Apply header styles
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Populate the rows with data
    for row in df.itertuples(index=False, name=None):
        sheet.append(row)
        for cell in sheet[sheet.max_row]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Apply styling to semester rows for clear separation
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value and not row[1].value:  
            for cell in row:
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook to a BytesIO stream
    excel_output = io.BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)
    return excel_output

# Function to plot GPA progress 
def plot_gpa_progress(gpas):
    if gpas: 
        fig, ax = plt.subplots(figsize=(10, 6))  

        semesters = list(range(1, len(gpas) + 1))
        
        semesters.insert(0, 0)
        gpas.insert(0, 0)

        ax.plot(semesters, gpas, marker='o', linestyle='-', color='b', label="GPA Progress")

        ax.set_ylim(0, max(gpas) + 1)  

        ax.set_xticks(range(0, len(gpas) + 1)) 
        ax.set_xticklabels(range(0, len(gpas) + 1))  

        ax.set_yticks(range(0, int(max(gpas)) + 2))  
        ax.set_yticklabels(range(0, int(max(gpas)) + 2))  

        ax.yaxis.set_major_locator(MaxNLocator(integer=True, prune='lower'))  

        ax.set_title("GPA Progress Over Semesters", fontsize=14)
        ax.set_xlabel("Semester", fontsize=12)
        ax.set_ylabel("GPA", fontsize=12)
        
        ax.grid(True)
        
        ax.legend()

        st.pyplot(fig)

# Streamlit App Logic
st.title("GPA & CGPA Calculator")

if "data" not in st.session_state:
    st.session_state["data"] = []
    st.session_state["gpas"] = []
    st.session_state["credits"] = []

total_semesters = st.number_input("Enter total number of semesters:", min_value=1, step=1)
completed_semesters = st.number_input("Enter number of semesters completed:", min_value=0, max_value=total_semesters, step=1)

new_data = []

for semester in range(1, completed_semesters + 1):
    st.subheader(f"Semester {semester}")
    num_subjects = st.number_input(f"Enter number of subjects for Semester {semester}:", min_value=1, step=1, key=f"subjects_{semester}")

    subjects = []
    marks = []
    credits = []

    for i in range(num_subjects):
        subject_name = st.text_input(f"Enter subject name for subject {i+1}:", key=f"subject_name_{semester}_{i}")
        mark = st.number_input(f"Enter marks for {subject_name}:", min_value=0, max_value=100, key=f"marks_{semester}_{i}")
        credit = st.number_input(f"Enter credits for {subject_name}:", min_value=0, key=f"credit_{semester}_{i}")

        subjects.append(subject_name)
        marks.append(mark)
        credits.append(credit)

    gpa = calculate_gpa(marks, credits)
    new_data.append({
        "semester": semester,
        "subjects": subjects,
        "marks": marks,
        "credits": credits,
        "gpa": gpa
    })

if st.button("Calculate"):
    st.session_state["data"] = new_data
    st.session_state["gpas"] = [semester_data["gpa"] for semester_data in new_data]
    st.session_state["credits"] = [sum(semester_data["credits"]) for semester_data in new_data]

    st.success("GPA and CGPA calculated!")

# Display GPA and CGPA
if st.session_state["data"]:
    for semester_data in st.session_state["data"]:
        st.write(f"Semester {semester_data['semester']} GPA: {semester_data['gpa']:.2f}")

    cgpa = calculate_cgpa(st.session_state["gpas"], st.session_state["credits"])
    st.write(f"Overall CGPA: {cgpa:.2f}")

    # Plot GPA Progress
    plot_gpa_progress(st.session_state["gpas"])

    # Generate Reports
    if st.button("Download PDF Report"):
        pdf_file = generate_pdf(st.session_state["data"], cgpa)
        st.download_button("Download PDF", data=pdf_file, file_name="gpa_report.pdf", mime="application/pdf")

    if st.button("Download Excel Report"):
        excel_file = generate_excel(st.session_state["data"], cgpa)
        st.download_button("Download Excel", data=excel_file, file_name="gpa_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
